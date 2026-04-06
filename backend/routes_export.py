"""
Routes para Exportación a Excel
Exporta documentos, tareas, empresas a Excel
"""

from flask import Blueprint, send_file, request
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from models import Documento, Tarea, Empresa, User, db
from tenant_utils import get_current_gestoria_id
from constants import DocumentCategories, TaskStates
from decorators_rbac import super_admin_required
from formatters import DateFormatter, CurrencyFormatter

export_bp = Blueprint('export', __name__)


def crear_estilo_header():
    """Estilo para headers de Excel"""
    return {
        'font': Font(bold=True, color="FFFFFF", size=12),
        'fill': PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def aplicar_estilo_header(ws, row=1):
    """Aplica estilo a la fila de headers"""
    estilo = crear_estilo_header()
    for cell in ws[row]:
        cell.font = estilo['font']
        cell.fill = estilo['fill']
        cell.alignment = estilo['alignment']
        cell.border = estilo['border']


def ajustar_anchos_columnas(ws):
    """Ajusta el ancho de las columnas automáticamente"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ============================================
# EXPORTACIÓN PARA USUARIOS NORMALES
# ============================================

@export_bp.route('/api/export/documentos', methods=['GET'])
@login_required
def export_documentos():
    """
    Exporta documentos a Excel
    Filtrado por gestoría del usuario actual
    """
    try:
        gestoria_id = get_current_gestoria_id()
        print(f"📥 Exportando documentos para gestoría {gestoria_id}")
        
        # Obtener documentos
        documentos = Documento.query.filter(
            Documento.gestoria_id == gestoria_id
        ).order_by(Documento.fecha_creacion.desc()).all()
        
        print(f"📊 Total documentos encontrados: {len(documentos)}")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Documentos"
        
        # Headers
        headers = [
            'ID', 'Nombre Archivo', 'Empresa', 'Categoría',
            'Fecha Creación', 'Fecha Procesado', 'Estado'
        ]
        ws.append(headers)
        aplicar_estilo_header(ws)
        
        # Datos
        for doc in documentos:
            try:
                ws.append([
                    doc.id,
                    doc.nombre_archivo,
                    doc.empresa.nombre if doc.empresa else 'Sin empresa',
                    doc.categoria,
                    DateFormatter.format_display(doc.fecha_creacion) if doc.fecha_creacion else '',
                    DateFormatter.format_display(doc.fecha_procesado) if doc.fecha_procesado else 'Pendiente',
                    'Procesado' if doc.fecha_procesado else 'Pendiente'
                ])
            except Exception as e:
                print(f"⚠️ Error procesando documento {doc.id}: {e}")
                continue
        
        # Ajustar anchos
        ajustar_anchos_columnas(ws)
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'documentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        print(f"✅ Excel generado: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando documentos: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500


@export_bp.route('/api/export/tareas', methods=['GET'])
@login_required
def export_tareas():
    """
    Exporta tareas a Excel
    Filtrado por gestoría del usuario actual
    """
    try:
        gestoria_id = get_current_gestoria_id()
        print(f"📥 Exportando tareas para gestoría {gestoria_id}")
        
        # Obtener tareas filtradas por gestoria_id
        # NOTA: Requiere que la tabla tareas tenga gestoria_id
        # Si no existe, ejecutar: migrations/add_gestoria_to_tareas.sql
        tareas = Tarea.query.filter(
            Tarea.gestoria_id == gestoria_id
        ).order_by(Tarea.fecha_creacion.desc()).all()
        
        print(f"📊 Total tareas encontradas: {len(tareas)}")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tareas"
        
        # Headers
        headers = [
            'ID', 'Título', 'Descripción', 'Empresa',
            'Estado', 'Prioridad', 'Asignado a',
            'Fecha Creación', 'Fecha Vencimiento', 'Fecha Completado',
            'Origen', 'Creado Por', 'Notas', 'Tags'
        ]
        ws.append(headers)
        aplicar_estilo_header(ws)
        
        # Datos
        for tarea in tareas:
            try:
                # Formatear tags como string
                tags_str = ', '.join(tarea.tags) if tarea.tags and isinstance(tarea.tags, list) else ''
                
                ws.append([
                    tarea.id,
                    tarea.titulo,
                    tarea.descripcion or '',
                    tarea.empresa.nombre if tarea.empresa else '',
                    tarea.estado,
                    tarea.prioridad or 'media',
                    tarea.asignado_a.nombre if hasattr(tarea, 'asignado_a') and tarea.asignado_a else 'Sin asignar',
                    DateFormatter.format_display(tarea.fecha_creacion) if tarea.fecha_creacion else '',
                    DateFormatter.format_display(tarea.fecha_vencimiento) if hasattr(tarea, 'fecha_vencimiento') and tarea.fecha_vencimiento else '',
                    DateFormatter.format_display(tarea.fecha_completada) if hasattr(tarea, 'fecha_completada') and tarea.fecha_completada else '',
                    # Nuevos campos de tracking
                    tarea.origen or 'manual',
                    tarea.creado_por.nombre if hasattr(tarea, 'creado_por') and tarea.creado_por else '',
                    tarea.notas or '',
                    tags_str
                ])
            except Exception as e:
                print(f"⚠️ Error procesando tarea {tarea.id}: {e}")
                continue
        
        # Ajustar anchos
        ajustar_anchos_columnas(ws)
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        print(f"✅ Excel generado: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando tareas: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500


@export_bp.route('/api/export/empresas', methods=['GET'])
@login_required
def export_empresas():
    """
    Exporta empresas a Excel
    Filtrado por gestoría del usuario actual
    """
    try:
        gestoria_id = get_current_gestoria_id()
        print(f"📥 Exportando empresas para gestoría {gestoria_id}")
        
        # Obtener empresas con stats
        from sqlalchemy import func
        
        empresas = db.session.query(
            Empresa,
            func.count(Documento.id).label('total_docs')
        ).outerjoin(Documento).filter(
            Empresa.gestoria_id == gestoria_id
        ).group_by(Empresa.id).all()
        
        print(f"📊 Total empresas encontradas: {len(empresas)}")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"
        
        # Headers (19 campos solicitados + Stats)
        headers = [
            'Codigo Empresa', 'NIF-NIE-CIF', 'Nombre Sociedad', 'Email', 'Telefono',
            'Nombre Administrador', 'Apellido Administrador', 'NIF-NIE-CIF ADMINISTRADOR',
            'Provincia', 'Municipio', 'Código Postal', 'Dirección', 'Dirección Centros Trabajo',
            'Cuenta Cotizacion', 'Convenio Colectivo Número', 'Convenio Colectivo Nombre',
            'EPIGRAFE IAE', 'CNAE 2009', 'CNAE 2025',
            'ID Registro', 'Estado', 'Total Documentos', 'Fecha Creación'
        ]
        ws.append(headers)
        aplicar_estilo_header(ws)
        
        # Datos
        for empresa, total_docs in empresas:
            try:
                ws.append([
                    empresa.codigo_empresa or '',
                    empresa.nif or '',
                    empresa.nombre or '',
                    empresa.email or '',
                    empresa.telefono or '',
                    empresa.nombre_administrador or '',
                    empresa.apellido_administrador or '',
                    empresa.nif_administrador or '',
                    empresa.provincia or '',
                    empresa.municipio or '',
                    empresa.codigo_postal or '',
                    empresa.direccion or '',
                    empresa.direccion_centros_trabajo_str or '',
                    empresa.cuenta_cotizacion or '',
                    empresa.convenio_numero or '',
                    empresa.convenio_nombre or '',
                    empresa.epigrafe_iae_str or '',
                    empresa.cnae_2009_str or '',
                    empresa.cnae_2025_str or '',
                    empresa.id,
                    empresa.estado or 'Abierto',
                    total_docs,
                    DateFormatter.format_display(empresa.fecha_creacion) if hasattr(empresa, 'fecha_creacion') and empresa.fecha_creacion else ''
                ])
            except Exception as e:
                print(f"⚠️ Error procesando empresa {empresa.id}: {e}")
                continue
        
        # Ajustar anchos
        ajustar_anchos_columnas(ws)
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'empresas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        print(f"✅ Excel generado: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando empresas: {e}")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500


# ============================================
# EXPORTACIÓN PARA SUPERADMIN
# ============================================

@export_bp.route('/api/admin/export/global', methods=['GET'])
@login_required
@super_admin_required
def export_global():
    """
    Exporta datos globales de todas las gestorías
    Solo para superadmin
    """
    try:
        from models import Gestoria
        
        # Crear workbook con múltiples hojas
        wb = Workbook()
        
        # Hoja 1: Gestorías
        ws_gestorias = wb.active
        ws_gestorias.title = "Gestorías"
        
        headers_gestorias = ['ID', 'Nombre', 'Total Empresas', 'Total Usuarios', 'Total Documentos']
        ws_gestorias.append(headers_gestorias)
        aplicar_estilo_header(ws_gestorias)
        
        gestorias = db.session.query(
            Gestoria,
            func.count(func.distinct(Empresa.id)).label('total_empresas'),
            func.count(func.distinct(User.id)).label('total_usuarios'),
            func.count(Documento.id).label('total_docs')
        ).outerjoin(Empresa, Gestoria.id == Empresa.gestoria_id
        ).outerjoin(User, Gestoria.id == User.gestoria_id
        ).outerjoin(Documento, Empresa.id == Documento.empresa_id
        ).group_by(Gestoria.id).all()
        
        for gestoria, empresas, usuarios, docs in gestorias:
            ws_gestorias.append([
                gestoria.id,
                gestoria.nombre,
                empresas,
                usuarios,
                docs
            ])
        
        ajustar_anchos_columnas(ws_gestorias)
        
        # Hoja 2: Todos los documentos
        ws_docs = wb.create_sheet("Documentos")
        
        headers_docs = ['ID', 'Nombre', 'Gestoría', 'Empresa', 'Categoría', 'Fecha Procesado']
        ws_docs.append(headers_docs)
        aplicar_estilo_header(ws_docs)
        
        documentos = db.session.query(
            Documento, Gestoria.nombre
        ).join(Empresa).join(Gestoria).limit(10000).all()  # Límite para no saturar
        
        for doc, gestoria_nombre in documentos:
            ws_docs.append([
                doc.id,
                doc.nombre_archivo,
                gestoria_nombre,
                doc.empresa.nombre if doc.empresa else '',
                doc.categoria,
                DateFormatter.format_display(doc.fecha_procesado) if doc.fecha_procesado else 'Pendiente'
            ])
        
        ajustar_anchos_columnas(ws_docs)
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'reporte_global_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exportando global: {e}")
        return {'error': str(e)}, 500


# ============================================
# NUEVAS EXPORTACIONES - GRUPOS Y FILTROS
# ============================================

@export_bp.route('/api/grupos-documentos/<int:grupo_id>/download-zip', methods=['GET'])
@login_required
def download_grupo_zip(grupo_id):
    """
    Descarga todos los documentos de un grupo en formato ZIP
    """
    try:
        from services.export_service import ExportService
        from models import GrupoDocumentos
        
        # Verificar que el grupo existe
        grupo = GrupoDocumentos.query.get(grupo_id)
        if not grupo:
            return {'error': 'Grupo no encontrado'}, 404
        
        # Generar ZIP
        zip_bytes, filename = ExportService.export_grupo_zip(grupo_id)
        
        # Enviar archivo
        output = BytesIO(zip_bytes)
        return send_file(
            output,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
    
    except ValueError as e:
        return {'error': str(e)}, 400
    except Exception as e:
        print(f"Error al generar ZIP: {e}")
        import traceback
        traceback.print_exc()
        return {'error': 'Error al generar archivo ZIP'}, 500


@export_bp.route('/api/export/documentos-filtrado', methods=['POST'])
@login_required
def export_documentos_filtrado():
    """
    Exporta documentos con filtros en el formato especificado
    
    Body:
        format: 'excel' | 'csv' | 'pdf'
        filters: {
            empresa_ids: [1, 2, 3],
            fecha_desde: '2025-01-01',
            fecha_hasta: '2025-12-31',
            categorias: ['nomina', 'notificacion']
        }
    """
    try:
        from services.export_service import ExportService
        
        data = request.get_json()
        export_format = data.get('format', 'excel')
        filters = data.get('filters', {})
        
        # Convertir fechas de string a datetime si existen
        if filters.get('fecha_desde'):
            filters['fecha_desde'] = datetime.strptime(filters['fecha_desde'], '%Y-%m-%d')
        if filters.get('fecha_hasta'):
            filters['fecha_hasta'] = datetime.strptime(filters['fecha_hasta'], '%Y-%m-%d')
        
        # Generar exportación según formato
        if export_format == 'excel':
            file_bytes, filename = ExportService.export_documentos_excel(filters)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif export_format == 'csv':
            file_content, filename = ExportService.export_documentos_csv(filters)
            file_bytes = file_content.encode('utf-8')
            mimetype = 'text/csv'
        elif export_format == 'pdf':
            file_bytes, filename = ExportService.export_reporte_pdf(filters)
            mimetype = 'application/pdf'
        else:
            return {'error': 'Formato no soportado'}, 400
        
        # Enviar archivo
        output = BytesIO(file_bytes)
        return send_file(
            output,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error al exportar documentos: {e}")
        import traceback
        traceback.print_exc()
        return {'error': f'Error al exportar: {str(e)}'}, 500
