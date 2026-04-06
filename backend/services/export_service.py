"""
Servicio de Exportación Masiva
Maneja la generación de archivos ZIP, Excel, CSV y PDF
"""
import os
import io
import zipfile
from datetime import datetime
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from models import db, Documento, GrupoDocumentos, Empresa, GrupoDocumentosItem


class ExportService:
    """Servicio para exportación de documentos"""
    
    @staticmethod
    def export_grupo_zip(grupo_id):
        """
        Exporta todos los documentos de un grupo en un archivo ZIP
        
        Args:
            grupo_id: ID del grupo de documentos
            
        Returns:
            tuple: (bytes del ZIP, nombre del archivo)
        """
        grupo = GrupoDocumentos.query.get(grupo_id)
        if not grupo:
            raise ValueError(f"Grupo {grupo_id} no encontrado")
        
        # Obtener documentos del grupo a través de la tabla intermedia
        items = GrupoDocumentosItem.query.filter_by(grupo_id=grupo_id).all()
        documentos = [item.documento for item in items if item.documento]
        
        if not documentos:
            raise ValueError(f"El grupo '{grupo.nombre}' no tiene documentos")
        
        # Crear ZIP en memoria
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Organizar por categoría
            categorias = {}
            for doc in documentos:
                categoria = doc.categoria or 'Sin_Categoria'
                if categoria not in categorias:
                    categorias[categoria] = []
                categorias[categoria].append(doc)
            
            # Añadir documentos al ZIP
            for categoria, docs in categorias.items():
                for doc in docs:
                    if doc.ruta_archivo and os.path.exists(doc.ruta_archivo):
                        # Ruta dentro del ZIP: Categoria/nombre_archivo.pdf
                        archivo_nombre = os.path.basename(doc.ruta_archivo)
                        zip_path = f"{categoria}/{archivo_nombre}"
                        
                        zip_file.write(doc.ruta_archivo, zip_path)
        
        zip_buffer.seek(0)
        
        # Nombre del archivo
        empresa = Empresa.query.get(grupo.empresa_id)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{empresa.nombre}_{grupo.nombre}_{timestamp}.zip"
        
        # Sanitizar nombre de archivo
        filename = "".join(c for c in filename if c.isalnum() or c in ('_', '-', '.')).rstrip()
        
        return zip_buffer.getvalue(), filename
    
    @staticmethod
    def export_documentos_excel(filters):
        """
        Exporta documentos a Excel con múltiples hojas
        
        Args:
            filters: dict con filtros (empresa_ids, fecha_desde, fecha_hasta, categorias)
            
        Returns:
            tuple: (bytes del Excel, nombre del archivo)
        """
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen de Empresas
        ws_empresas = wb.active
        ws_empresas.title = "Empresas"
        
        # Headers
        headers_empresas = ['ID', 'Nombre', 'NIF', 'Total Documentos', 'Última Actualización']
        ws_empresas.append(headers_empresas)
        
        # Estilo de headers
        for cell in ws_empresas[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos de empresas
        query = Empresa.query
        if filters.get('empresa_ids'):
            query = query.filter(Empresa.id.in_(filters['empresa_ids']))
        
        empresas = query.all()
        for empresa in empresas:
            # Contar documentos de esta empresa
            total_docs = Documento.query.filter_by(empresa_id=empresa.id).count()
            
            ws_empresas.append([
                empresa.id,
                empresa.nombre,
                empresa.nif or 'N/A',
                total_docs,
                empresa.updated_at.strftime('%Y-%m-%d %H:%M') if hasattr(empresa, 'updated_at') and empresa.updated_at else 'N/A'
            ])
        
        # Hoja 2: Documentos
        ws_docs = wb.create_sheet("Documentos")
        
        headers_docs = ['ID', 'Empresa', 'Categoría', 'Nombre Archivo', 'Fecha Procesado', 
                       'Estado', 'Importe', 'Fecha Documento']
        ws_docs.append(headers_docs)
        
        # Estilo de headers
        for cell in ws_docs[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Query de documentos
        query = db.session.query(Documento, Empresa).join(Empresa)
        
        # Aplicar filtros solo si tienen valores
        if filters.get('empresa_ids') and len(filters['empresa_ids']) > 0:
            query = query.filter(Documento.empresa_id.in_(filters['empresa_ids']))
        if filters.get('fecha_desde'):
            query = query.filter(Documento.fecha_creacion >= filters['fecha_desde'])
        if filters.get('fecha_hasta'):
            query = query.filter(Documento.fecha_creacion <= filters['fecha_hasta'])
        if filters.get('categorias') and len(filters['categorias']) > 0:
            query = query.filter(Documento.categoria.in_(filters['categorias']))
        
        documentos = query.all()
        print(f"✅ Excel: Total documentos encontrados: {len(documentos)}")
        print(f"🔍 Filtros aplicados: {filters}")
        
        for doc, empresa in documentos:
            ws_docs.append([
                doc.id,
                empresa.nombre,
                doc.categoria or 'Sin categoría',
                doc.nombre_archivo,
                doc.fecha_procesado.strftime('%Y-%m-%d %H:%M') if doc.fecha_procesado else 'N/A',
                'Procesado' if doc.procesado else 'Pendiente',
                doc.importe or doc.importe_pagar or 'N/A',
                doc.fecha_creacion.strftime('%Y-%m-%d') if doc.fecha_creacion else 'N/A'
            ])
        
        # Hoja 3: Datos Extraídos
        ws_datos = wb.create_sheet("Datos Extraídos")
        
        headers_datos = ['Documento ID', 'Empresa', 'Campo', 'Valor']
        ws_datos.append(headers_datos)
        
        for cell in ws_datos[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos extraídos
        for doc, empresa in documentos:
            if doc.datos_extraidos:
                for campo, valor in doc.datos_extraidos.items():
                    ws_datos.append([
                        doc.id,
                        empresa.nombre,
                        campo,
                        str(valor)
                    ])
        
        # Ajustar anchos de columnas
        for ws in [ws_empresas, ws_docs, ws_datos]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Exportacion_IAGES_{timestamp}.xlsx"
        
        return excel_buffer.getvalue(), filename
    
    @staticmethod
    def export_documentos_csv(filters):
        """
        Exporta documentos a CSV
        
        Args:
            filters: dict con filtros
            
        Returns:
            tuple: (string CSV, nombre del archivo)
        """
        import csv
        
        # Query de documentos
        query = db.session.query(Documento, Empresa).join(Empresa)
        
        # Aplicar filtros solo si tienen valores
        if filters.get('empresa_ids') and len(filters['empresa_ids']) > 0:
            query = query.filter(Documento.empresa_id.in_(filters['empresa_ids']))
        if filters.get('fecha_desde'):
            query = query.filter(Documento.fecha_creacion >= filters['fecha_desde'])
        if filters.get('fecha_hasta'):
            query = query.filter(Documento.fecha_creacion <= filters['fecha_hasta'])
        if filters.get('categorias') and len(filters['categorias']) > 0:
            query = query.filter(Documento.categoria.in_(filters['categorias']))
        
        documentos = query.all()
        print(f"✅ CSV: Total documentos encontrados: {len(documentos)}")
        print(f"🔍 Filtros CSV: {filters}")
        
        # Crear CSV en memoria
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        
        # Headers
        writer.writerow(['ID', 'Empresa', 'NIF', 'Categoría', 'Nombre Archivo', 
                        'Fecha Procesado', 'Estado', 'Importe', 'Fecha Documento'])
        
        # Datos
        for doc, empresa in documentos:
            writer.writerow([
                doc.id,
                empresa.nombre,
                empresa.nif or 'N/A',
                doc.categoria or 'Sin categoría',
                doc.nombre_archivo,
                doc.fecha_procesado.strftime('%Y-%m-%d %H:%M') if doc.fecha_procesado else 'N/A',
                'Procesado' if doc.procesado else 'Pendiente',
                doc.importe or doc.importe_pagar or 'N/A',
                doc.fecha_creacion.strftime('%Y-%m-%d') if doc.fecha_creacion else 'N/A'
            ])
        
        # Nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Exportacion_IAGES_{timestamp}.csv"
        
        return csv_buffer.getvalue(), filename
    
    @staticmethod
    def export_reporte_pdf(filters):
        """
        Genera reporte PDF profesional mejorado
        
        Args:
            filters: dict con filtros
            
        Returns:
            tuple: (bytes del PDF, nombre del archivo)
        """
        from reportlab.lib.pagesizes import landscape
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer, 
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        # Estilos personalizados
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=20,
            alignment=1
        )
        
        elements = []
        
        # Header con estilo
        title = Paragraph("<b>IAGES</b> - Reporte de Documentos", title_style)
        elements.append(title)
        
        # Fecha de generación
        fecha_gen = Paragraph(
            f"Generado el {datetime.now().strftime('%d de %B de %Y a las %H:%M')}", 
            subtitle_style
        )
        elements.append(fecha_gen)
        elements.append(Spacer(1, 0.5*cm))
        
        # Query de documentos
        query = db.session.query(Documento, Empresa).join(Empresa)
        
        # Aplicar filtros
        if filters.get('empresa_ids') and len(filters['empresa_ids']) > 0:
            query = query.filter(Documento.empresa_id.in_(filters['empresa_ids']))
        if filters.get('fecha_desde'):
            query = query.filter(Documento.fecha_creacion >= filters['fecha_desde'])
        if filters.get('fecha_hasta'):
            query = query.filter(Documento.fecha_creacion <= filters['fecha_hasta'])
        if filters.get('categorias') and len(filters['categorias']) > 0:
            query = query.filter(Documento.categoria.in_(filters['categorias']))
        
        # Obtener documentos
        documentos = query.all()
        total_docs = len(documentos)
        
        # Calcular total de importes
        total_importe = sum(
            float(doc.importe or doc.importe_pagar or 0) 
            for doc, _ in documentos 
            if (doc.importe or doc.importe_pagar)
        )
        
        # Resumen con estilo
        resumen_data = [
            ['Total de Documentos', str(total_docs)],
            ['Total Importes', f"{total_importe:,.2f}€"]
        ]
        
        resumen_table = Table(resumen_data, colWidths=[5*cm, 4*cm])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#3b82f6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#eff6ff')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e40af')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('PADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ]))
        
        elements.append(resumen_table)
        elements.append(Spacer(1, 1*cm))
        
        if total_docs == 0:
            no_data = Paragraph(
                "<i>No se encontraron documentos con los filtros aplicados.</i>", 
                styles['Normal']
            )
            elements.append(no_data)
        else:
            # Tabla de documentos con mejor diseño
            data = [['Empresa', 'Categoría', 'Fecha', 'Importe (€)']]
            
            for documento, empresa in documentos:
                importe_val = float(documento.importe or documento.importe_pagar or 0)
                data.append([
                    Paragraph(empresa.nombre[:40] if empresa.nombre else 'N/A', styles['Normal']),
                    documento.categoria or 'Sin categoría',
                    documento.fecha_creacion.strftime('%d/%m/%Y') if documento.fecha_creacion else 'N/A',
                    f"{importe_val:,.2f}" if importe_val > 0 else '-'
                ])
            
            # Ajustar anchos de columna
            col_widths = [8*cm, 5*cm, 3*cm, 3*cm]
            table = Table(data, colWidths=col_widths, repeatRows=1)
            
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
                ('TOPPADDING', (0, 0), (-1, 0), 14),
                
                # Body
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('PADDING', (0, 1), (-1, -1), 8),
                
                # Alternating rows
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1e40af')),
            ]))
            
            elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Reporte_IAGES_{timestamp}.pdf"
        
        return pdf_buffer.getvalue(), filename
